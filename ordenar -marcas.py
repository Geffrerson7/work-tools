from openpyxl import load_workbook

def add_separator_between_values(input_file_path, output_file_path):
    # Cargar el archivo Excel
    wb = load_workbook(filename=input_file_path)
    
    # Seleccionar la hoja de cálculo
    sheet = wb.active
    
    # Obtener los valores de la columna "Marca"
    marca_column = sheet['A']  # Suponiendo que la columna "Marca" está en la columna A

    # Crear una lista para almacenar los nuevos valores con los separadores y fila vacía
    new_values = []

    # Iterar sobre los valores de la columna "Marca" y agregarlos a la lista con el separador "---"
    for value in marca_column:
        new_values.append(value.value)
        new_values.extend(['---'] * 4)  # Cuatro separadores

        # Agregar la quinta fila vacía
        new_values.append('')

    # Insertar las celdas con los valores en la hoja de cálculo
    for i, separator in enumerate(new_values, start=1):
        sheet.cell(row=i, column=1, value=separator)

    # Guardar los cambios en el nuevo archivo Excel
    wb.save(output_file_path)

# Rutas de los archivos de entrada y salida
input_file_path = "marcas-espe.xlsx"
output_file_path = "output.xlsx"

# Llamar a la función para agregar los separadores y la fila vacía entre los valores de la columna "Marca"
add_separator_between_values(input_file_path, output_file_path)
